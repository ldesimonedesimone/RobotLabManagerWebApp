#!/usr/bin/env python3
"""Parse schedule template XLSX and load into the schedule_templates table.

Usage (from backend/):
    .venv/bin/python parse_templates.py "../Schedule Templates (1).xlsx"

Reads SCHEDULE_DATABASE_URL from backend/.env.

XLSX convention — one template per sheet:
  - Sheet name = template name (sheets named "Sheet1" etc. auto-name as "5P-3R-2T").
  - Pilot legend: rows where col A = "Pilot 1", "Pilot 2", … and col B has a
    colored fill swatch (value ignored; only the background color matters).
  - Robot rows: col A = "Robot 1", "Robot 2", … Col B onward = color-coded cells.
  - Task rows : col A = "Task 1", "Task 2", … Col B onward = color-coded cells.
  - Blank rows between sections are fine; they are skipped.
  - Each colored cell in the grid maps to a pilot via the legend colors.
  - Each column in the grid = one 15-minute time slot.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import psycopg
from dotenv import dotenv_values

BACKEND_DIR = Path(__file__).resolve().parent


def get_db_url() -> str:
    env = dotenv_values(BACKEND_DIR / ".env")
    url = (env.get("SCHEDULE_DATABASE_URL") or "").strip()
    assert url, "SCHEDULE_DATABASE_URL not set in backend/.env"
    return url


def _cell_rgb(cell) -> str | None:
    fg = cell.fill.fgColor
    if fg and fg.rgb and str(fg.rgb) != "00000000":
        return str(fg.rgb)
    return None


def parse_sheet(ws) -> dict:
    pilot_colors: list[str] = []

    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip().lower().startswith("pilot"):
            rgb = _cell_rgb(ws.cell(r, 2))
            if rgb:
                pilot_colors.append(rgb)

    assert pilot_colors, f"No pilot legend found in sheet '{ws.title}'"

    color_to_idx: dict[str, int] = {c: i for i, c in enumerate(pilot_colors)}

    robot_rows: list[int] = []
    task_rows: list[int] = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if not val:
            continue
        s = str(val).strip().lower()
        if s.startswith("robot"):
            robot_rows.append(r)
        elif s.startswith("task"):
            task_rows.append(r)

    all_rows = robot_rows + task_rows
    assert all_rows, f"No Robot/Task rows found in sheet '{ws.title}'"

    grid_by_row: list[list[int | None]] = []
    for r in all_rows:
        row_data: list[int | None] = []
        for c in range(2, ws.max_column + 1):
            rgb = _cell_rgb(ws.cell(r, c))
            row_data.append(color_to_idx.get(rgb) if rgb else None)
        while row_data and row_data[-1] is None:
            row_data.pop()
        grid_by_row.append(row_data)

    n_slots = max((len(row) for row in grid_by_row), default=0)
    for row in grid_by_row:
        row.extend([None] * (n_slots - len(row)))

    n_resources = len(grid_by_row)
    grid: list[list[int | None]] = []
    for t in range(n_slots):
        grid.append([grid_by_row[ri][t] for ri in range(n_resources)])

    return {
        "n_pilots": len(pilot_colors),
        "n_robots": len(robot_rows),
        "n_tasks": len(task_rows),
        "grid": grid,
    }


def main() -> None:
    assert len(sys.argv) > 1, "Usage: .venv/bin/python parse_templates.py <path.xlsx>"
    xlsx_path = sys.argv[1]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    db_url = get_db_url()

    with psycopg.connect(db_url) as conn:
        for sheet_name in wb.sheetnames:
            tpl = parse_sheet(wb[sheet_name])
            name = sheet_name.strip()
            if name.lower().startswith("sheet"):
                name = f"{tpl['n_pilots']}P-{tpl['n_robots']}R-{tpl['n_tasks']}T"

            print(
                f"Parsed: {name}  "
                f"({tpl['n_pilots']}P {tpl['n_robots']}R {tpl['n_tasks']}T, "
                f"{len(tpl['grid'])} slots)"
            )

            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO schedule_templates (name, n_pilots, n_robots, n_tasks, grid)
                    VALUES (%s, %s, %s, %s, %s::jsonb)
                    ON CONFLICT (name) DO UPDATE SET
                        n_pilots = EXCLUDED.n_pilots,
                        n_robots = EXCLUDED.n_robots,
                        n_tasks = EXCLUDED.n_tasks,
                        grid = EXCLUDED.grid
                    """,
                    (name, tpl["n_pilots"], tpl["n_robots"], tpl["n_tasks"], json.dumps(tpl["grid"])),
                )
            conn.commit()
            print(f"  -> Upserted '{name}'")

    print("Done.")


if __name__ == "__main__":
    main()
